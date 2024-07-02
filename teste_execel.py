import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

url = "https://consultcpf-devaprender.netlify.app"

planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

driver = webdriver.Edge()
driver.get(url)
sleep(4) 

   
for linha in pagina_clientes.iter_rows(values_only=True,min_row=2,max_row=10):#quando vc coloca esse values_only não precisa colocar .value no final da linha[0].value
    nome = (str(linha[0]))
    valor = (str(linha[1]))
    cpf = (str(linha[2]))
    vencimento = (str(linha[3]))
    

    campo_pesquisa = driver.find_element(By.XPATH,'//*[@id="cpfInput"]')
    sleep(1)
    campo_pesquisa.clear()
    sleep(0.5)
    campo_pesquisa.send_keys(cpf)
    sleep(1)
    botao = driver.find_element(By.XPATH,'//*[@id="consultaForm"]/button')
    sleep(0.5)
    botao.click()
    sleep(4)
    status = driver.find_element(By.XPATH,'//*[@id="statusLabel"]')
    if status.text == 'em dia':
        sleep(1)
        data_pagamento = driver.find_element(By.XPATH,'//*[@id="paymentDate"]')
        metodo_pegamento = driver.find_element(By.XPATH,'//*[@id="paymentMethod"]')
        sleep(0.5)
        data_pagamento_limpo = data_pagamento.text.split()[3] #fatiando uma string e pegando so o final dela 
        metodo_pegamento_limpo = metodo_pegamento.text.split()[3]
        
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome,valor,cpf,vencimento,'em dia ', data_pagamento_limpo,metodo_pegamento_limpo])
        
        planilha_fechamento.save('planilha fechamento.xlsx')
    else:
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome,valor,cpf,vencimento,'pendente'])
        
        planilha_fechamento.save('planilha fechamento.xlsx')
